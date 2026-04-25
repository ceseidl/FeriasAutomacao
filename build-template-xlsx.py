"""
Gera Ferias-template.xlsx a partir do antigo ferias-2026.xlsx, adicionando
sheets de referencia (Squads, Integrantes, Status) e dropdowns (data
validation) nas colunas Colaborador, Squad e Status da aba Ferias.

Como funciona:
  - Le ferias-2026.xlsx (ou Ferias-template.xlsx ja existente).
  - Extrai squads e integrantes distintos da aba Ferias.
  - Cria/atualiza as 3 sheets de lookup como Tabelas Excel (ListObject),
    pra que o usuario possa adicionar linhas no fim sem se preocupar.
  - Adiciona data validation tipo list nas colunas relevantes.
  - Reescreve a aba Instrucoes pra documentar o novo fluxo.
  - Salva como Ferias-template.xlsx.

Idempotente: se Ferias-template.xlsx ja existir, atualiza in-place sem
duplicar Tabelas ou data validations.

Uso:
    python build-template-xlsx.py
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent
NEW = ROOT / 'Ferias-template.xlsx'
OLD = ROOT / 'ferias-2026.xlsx'

if NEW.exists():
    src = NEW
    print(f'Usando {NEW.name} como base (atualizacao in-place).')
elif OLD.exists():
    src = OLD
    print(f'Usando {OLD.name} como base (primeira geracao).')
else:
    print(f'ERRO: nenhuma planilha base encontrada.', file=sys.stderr)
    sys.exit(1)

wb = load_workbook(src)
ferias = wb['Ferias']

# --- Extrai dados distintos da aba Ferias ----
# colunas: A=Mes, B=Colaborador, C=Squad, D=Inicio, E=Fim, F=Dias, G=Status
squads_set = set()
integrantes_set = set()  # (nome, squad)
for row in ferias.iter_rows(min_row=2, values_only=True):
    if row[2]:
        squads_set.add(str(row[2]).strip())
    if row[1]:
        integrantes_set.add((str(row[1]).strip(), str(row[2]).strip() if row[2] else ''))

squads = sorted(squads_set)
integrantes = sorted(integrantes_set, key=lambda t: t[0].lower())
statuses = ['Aprovada', 'Planejada', 'Solicitada']

print(f'  Squads distintos:      {len(squads)}')
print(f'  Integrantes distintos: {len(integrantes)}')

# --- helpers ----
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='305496')


def reset_sheet(name):
    """Cria a sheet do zero (apaga se ja existe). Retorna a worksheet."""
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def write_table(ws, headers, rows, table_name, style='TableStyleMedium9'):
    """Escreve cabecalho + dados e marca como Tabela Excel."""
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='left', vertical='center')
    for ri, row in enumerate(rows, start=2):
        if isinstance(row, (list, tuple)):
            for ci, v in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value=v)
        else:
            ws.cell(row=ri, column=1, value=row)
    last_row = max(2, 1 + len(rows))
    last_col_letter = chr(ord('A') + len(headers) - 1)
    ref = f'A1:{last_col_letter}{last_row}'
    # se nao houver dados, reserva uma linha vazia pra Tabela ser valida
    if not rows:
        ws.cell(row=2, column=1, value=None)
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    # ajustes de largura
    for ci, h in enumerate(headers, start=1):
        col_letter = chr(ord('A') + ci - 1)
        max_len = max([len(h)] + [len(str(r[ci-1] if isinstance(r, (list,tuple)) else r)) for r in rows] or [10])
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


# --- 1) Squads ----
ws = reset_sheet('Squads')
write_table(ws, ['Squad'], squads, 'tblSquads')

# --- 2) Integrantes ----
ws = reset_sheet('Integrantes')
write_table(ws, ['Integrante', 'Squad'], integrantes, 'tblIntegrantes')

# --- 3) Status ----
ws = reset_sheet('Status')
write_table(ws, ['Status'], statuses, 'tblStatus')

# --- 4) Data validation na aba Ferias ----
# Limpa data validations antigas pra nao duplicar
ferias.data_validations.dataValidation = []

# Squads -> coluna C (Squad)
dv_sq = DataValidation(
    type='list', formula1=f"=Squads!$A$2:$A${1 + max(len(squads), 1)}",
    allow_blank=True, showErrorMessage=True,
)
dv_sq.error = 'Squad fora da lista. Adicione na aba "Squads" antes de usar.'
dv_sq.errorTitle = 'Squad invalido'
dv_sq.prompt = 'Escolha um squad da lista'
dv_sq.promptTitle = 'Squad'
dv_sq.add('C2:C1000')
ferias.add_data_validation(dv_sq)

# Integrantes -> coluna B (Colaborador)
dv_int = DataValidation(
    type='list', formula1=f"=Integrantes!$A$2:$A${1 + max(len(integrantes), 1)}",
    allow_blank=True, showErrorMessage=True,
)
dv_int.error = 'Integrante fora da lista. Adicione na aba "Integrantes" antes de usar.'
dv_int.errorTitle = 'Integrante invalido'
dv_int.prompt = 'Escolha um integrante da lista'
dv_int.promptTitle = 'Colaborador'
dv_int.add('B2:B1000')
ferias.add_data_validation(dv_int)

# Status -> coluna G (Status)
dv_st = DataValidation(
    type='list', formula1=f"=Status!$A$2:$A${1 + len(statuses)}",
    allow_blank=True, showErrorMessage=True,
)
dv_st.error = 'Status fora da lista. Use Aprovada, Planejada ou Solicitada.'
dv_st.errorTitle = 'Status invalido'
dv_st.prompt = 'Escolha o status da ferias'
dv_st.promptTitle = 'Status'
dv_st.add('G2:G1000')
ferias.add_data_validation(dv_st)

# --- 5) Reescreve aba Instrucoes ----
ws_inst = reset_sheet('Instrucoes')
linhas = [
    'Instrucoes de uso',
    '',
    "1. A aba 'Ferias' e a planilha principal. Cada linha = um periodo de ferias de um colaborador.",
    "2. Colunas obrigatorias (nao renomeie): Mes, Colaborador, Squad, Inicio, Fim, Dias, Status.",
    "3. Colaborador, Squad e Status tem dropdowns - eles puxam das abas auxiliares:",
    "     - Squads:      lista de squads disponiveis (use a aba 'Squads' pra adicionar).",
    "     - Integrantes: lista de pessoas (use a aba 'Integrantes', com nome + squad).",
    "     - Status:      Aprovada, Planejada ou Solicitada (aba 'Status').",
    "4. Pra adicionar um squad/integrante novo, ir na aba correspondente e colar/digitar",
    "   no fim da tabela. O dropdown puxa automaticamente.",
    "5. Mes: use Janeiro, Fevereiro, Março, Abril, Maio, Junho, Julho, Agosto, Setembro,",
    "   Outubro, Novembro, Dezembro.",
    "6. Inicio/Fim: formato dd/mm/aaaa (ex: 12/01/2026).",
    "7. Dias: numero inteiro.",
    "",
    'Apos editar, salve o arquivo e gere o relatorio pelo atalho "Gerar Relatorio".',
]
for i, txt in enumerate(linhas, start=1):
    c = ws_inst.cell(row=i, column=1, value=txt)
    if i == 1:
        c.font = Font(bold=True, size=14)
ws_inst.column_dimensions['A'].width = 100

# --- 6) Reordena as sheets pra ficarem numa ordem logica ----
ordem_desejada = ['Ferias', 'Squads', 'Integrantes', 'Status', 'Instrucoes']
wb._sheets = [wb[name] for name in ordem_desejada if name in wb.sheetnames]

# --- 7) Salva ----
wb.save(NEW)
print(f'OK: {NEW.name} salvo.')
print(f'  Sheets: {wb.sheetnames}')
