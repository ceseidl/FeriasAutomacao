"""
Popula a aba 'Integrantes' do Ferias-template.xlsx com datas de exemplo
(admissao + ultimas ferias) pros 13 integrantes ja cadastrados.

Os valores sao escolhidos pra exercitar todos os blocos da nova secao
"Controle de Vencimento de Ferias" no relatorio gerado:

  - 2 pessoas no bloco CRITICO em dobro (2o vencimento ja passou)
  - 1 pessoa no bloco CRITICO proximo (2o vencimento dentro de 6 meses)
  - 3 pessoas no bloco ATENCAO (1o vencimento atingido ou proximo)
  - 5 pessoas em dia (nao aparecem na secao)
  - 2 pessoas com dados incompletos (sem data de admissao)

Considerando "hoje" = 25/04/2026.

Uso:
    python docs/populate-integrantes-example.py
"""
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'Ferias-template.xlsx'

# Mapa: nome -> (data_admissao, data_fim_ultimas_ferias)
# None = celula fica vazia
DADOS = {
    'Alan Antonio Pasini':         (date(2024,  3, 15), None),                  # CRITICO em dobro
    'Albert Samuel Melo':          (date(2024,  6, 20), None),                  # CRITICO proximo
    'Amanda Lobato Sampaio':       (date(2024,  9, 12), None),                  # ATENCAO (1o passou)
    'Andre Luis Borges':           (date(2025,  8,  3), None),                  # ATENCAO proximo
    'Carlos Seidl':                (date(2024,  4,  8), date(2025, 3, 15)),     # ATENCAO (1o passou)
    'Cibele Florido Siqueira':     (date(2026,  1, 10), None),                  # OK (recente)
    'Eduardo Cavalcante de Melo':  (date(2026,  2, 22), None),                  # OK (recente)
    'Eduardo Barbosa de Azevedo':  (date(2024,  6, 18), date(2026, 2, 25)),     # OK (tirou ferias recente)
    'Joao Pedro Ferreira Carvalho':(None,               None),                  # DADOS INCOMPLETOS
    'Joao Vitor Araujo Moura':     (date(2024,  5, 12), date(2026, 3, 30)),     # OK (tirou ferias recente)
    'Marcus Vinicius Andrade Luz': (date(2026,  1, 14), None),                  # OK (recente)
    'Paola Rosa de Lima Martins':  (date(2024,  2, 10), None),                  # CRITICO em dobro
    'Vagner dos Santos Behs':      (None,               None),                  # DADOS INCOMPLETOS
}

wb = load_workbook(XLSX)
ws = wb['Integrantes']

# Header esta na linha 1: A=Integrante, B=Squad, C=Data de inicio na AIR, D=Data das ultimas ferias
nao_encontrados = []
preenchidos = 0
for row in range(2, ws.max_row + 1):
    nome = ws.cell(row=row, column=1).value
    if not nome:
        continue
    nome = str(nome).strip()
    if nome not in DADOS:
        nao_encontrados.append(nome)
        continue
    adm, ult = DADOS[nome]
    ws.cell(row=row, column=3).value = adm
    ws.cell(row=row, column=4).value = ult
    ws.cell(row=row, column=3).number_format = 'dd/mm/yyyy'
    ws.cell(row=row, column=4).number_format = 'dd/mm/yyyy'
    preenchidos += 1

if nao_encontrados:
    print(f'AVISO: {len(nao_encontrados)} nome(s) nao encontrado(s) no DADOS:')
    for n in nao_encontrados:
        print(f'  - {n}')

wb.save(XLSX)
print(f'OK: {preenchidos} integrante(s) populado(s) com datas de exemplo em {XLSX.name}')
